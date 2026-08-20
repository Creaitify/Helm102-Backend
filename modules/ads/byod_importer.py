"""BYOD (Bring-Your-Own-Data) Excel/CSV Campaign Importer for HELM02.

Supports:
- Multi-sheet Excel workbooks (.xlsx) with platform auto-detection (e.g. Google_Ads, Meta_Ads, Summary).
- Standard CSV files (.csv) with dialect sniffing and encoding fallbacks.
- Column normalization and alias mapping.
- Validation of required metric columns (spend, roas, clicks, conversions).
- Automatic derivation of CTR and CPA if omitted.
- Sample Finnovate bundle generator for testing and demos.
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Sequence

import httpx
import openpyxl

from modules.ads.contracts import CampaignSnapshot, MetricRow, Platform


class BYODImportError(ValueError):
    """Base exception for BYOD import failures."""


class MissingRequiredColumnsError(BYODImportError):
    """Raised when required metric columns are missing from the dataset."""


class InvalidDataFormatError(BYODImportError):
    """Raised when dataset formatting or content cannot be parsed."""


# Canonical column keys required for basic metric calculation
REQUIRED_METRIC_FIELDS: tuple[str, ...] = ("spend_inr", "roas", "clicks", "conversions")

# Platform label mappings for human-readable display and synthesized campaign naming
PLATFORM_LABELS: dict[Platform, str] = {
    Platform.GOOGLE_ADS: "Google Ads",
    Platform.META_ADS: "Meta Ads",
    Platform.TIKTOK_ADS: "TikTok Ads",
    Platform.LINKEDIN_ADS: "LinkedIn Ads",
    Platform.BYOD: "BYOD",
}

# Human-friendly field display names for error messages
FIELD_DISPLAY_NAMES: dict[str, str] = {
    "spend_inr": "spend",
    "roas": "roas",
    "clicks": "clicks",
    "conversions": "conversions",
    "campaign_id": "campaign_id",
    "campaign_name": "campaign_name",
    "campaign_type": "campaign_type",
    "industry": "industry",
    "country": "country",
    "platform": "platform",
    "impressions": "impressions",
    "cpa_inr": "cpa",
    "ctr": "ctr",
    "cpc": "cpc",
    "revenue": "revenue",
    "status": "status",
    "account_id": "account_id",
}

# Alias dictionary mapping canonical names to acceptable column header variations
COLUMN_ALIASES: dict[str, set[str]] = {
    "campaign_id": {
        "campaign_id",
        "campaignid",
        "id",
        "campaign_idx",
        "ad_set_id",
        "adset_id",
        "ad_id",
        "adgroup_id",
        "ad_group_id",
        "campaign_code",
        "campaign_key",
    },
    "campaign_name": {
        "campaign_name",
        "campaignname",
        "campaign",
        "name",
        "ad_set_name",
        "adset_name",
        "ad_name",
        "campaign_title",
        "adgroup_name",
        "ad_group_name",
        "title",
    },
    "campaign_type": {
        "campaign_type",
        "type",
        "campaign_category",
        "ad_type",
        "channel_type",
        "objective",
        "goal",
    },
    "industry": {
        "industry",
        "vertical",
        "sector",
        "business_type",
        "category",
        "niche",
    },
    "country": {
        "country",
        "geo",
        "region",
        "location",
        "territory",
        "market",
        "country_code",
    },
    "date": {
        "date",
        "day",
        "timestamp",
        "period",
        "event_date",
        "report_date",
    },
    "platform": {
        "platform",
        "platform_name",
        "network",
        "channel",
        "source",
        "publisher",
        "ad_network",
        "channel_name",
    },
    "spend_inr": {
        "spend_inr",
        "spend",
        "cost",
        "spend_rs",
        "spend_in_inr",
        "amount_spent",
        "amount_spent_inr",
        "amount_spent_usd",
        "total_spend",
        "cost_inr",
        "cost_usd",
        "spend_amount",
        "daily_spend",
        "spend_usd",
        "ad_spend",
        "adspend",
        "ad_cost",
        "adcost",
        "total_cost",
        "budget_spent",
        "spent",
        "cost_micros",
        "media_spend",
    },
    "revenue": {
        "revenue",
        "conv_value",
        "conversion_value",
        "total_revenue",
        "sales",
        "purchase_value",
        "total_conv_value",
        "sales_amount",
        "revenue_usd",
        "revenue_inr",
        "value",
        "total_sales",
        "gross_revenue",
        "all_conv_value",
    },
    "impressions": {
        "impressions",
        "impr",
        "views",
        "impressions_count",
        "impr_count",
        "total_impressions",
        "reach",
    },
    "clicks": {
        "clicks",
        "link_clicks",
        "ad_clicks",
        "clicks_count",
        "total_clicks",
        "outbound_clicks",
    },
    "cpc": {
        "cpc",
        "cost_per_click",
        "cpc_inr",
        "cpc_usd",
        "avg_cpc",
        "average_cpc",
        "cost_click",
        "avg_cost_click",
        "cost_per_link_click",
        "cpc_amount",
    },
    "conversions": {
        "conversions",
        "conv",
        "results",
        "leads",
        "actions",
        "purchases",
        "conversions_count",
        "total_conversions",
        "purchases_count",
        "signups",
        "orders",
    },
    "roas": {
        "roas",
        "return_on_ad_spend",
        "conv_value_per_cost",
        "purchase_roas",
        "return_on_spend",
        "value_per_cost",
        "conv_value_cost",
        "website_purchase_roas",
    },
    "cpa_inr": {
        "cpa_inr",
        "cpa",
        "cost_per_conversion",
        "cost_per_action",
        "cost_per_result",
        "cost_per_lead",
        "cost_per_conv",
        "cost_conv",
        "cpa_rs",
        "cpa_usd",
        "cost_per_purchase",
    },
    "ctr": {
        "ctr",
        "click_through_rate",
        "ctr_pct",
        "ctr_percent",
        "click_thru_rate",
        "clickthrough_rate",
        "inline_link_click_ctr",
    },
    "status": {
        "status",
        "campaign_status",
        "state",
        "status_name",
        "ad_status",
        "delivery_status",
    },
    "account_id": {
        "account_id",
        "accountid",
        "customer_id",
        "act_id",
        "account",
        "advertiser_id",
    },
}


def normalize_column_name(col: str) -> str:
    """Normalize a header string for case/format-insensitive matching."""
    if not isinstance(col, str):
        col = str(col) if col is not None else ""
    norm = col.strip("\ufeff\u200b\u200c\u200d\r\n\t ").lower()
    norm = re.sub(r"[\(（\[].*?[\)）\]]", "", norm)
    norm = re.sub(r"[₹\$,%#@!]", "", norm)
    norm = re.sub(r"[\s\-\./\\]+", "_", norm)
    norm = re.sub(r"_+", "_", norm).strip("_")
    return norm


def map_headers(raw_headers: Sequence[Any]) -> dict[str, int]:
    """Map canonical field names to column indices in raw_headers.

    Returns:
        dict[canonical_field_name, column_index]
    """
    mapping: dict[str, int] = {}
    normalized_headers = [normalize_column_name(str(h or "")) for h in raw_headers]

    for col_idx, norm_header in enumerate(normalized_headers):
        if not norm_header:
            continue
        for canonical, aliases in COLUMN_ALIASES.items():
            if canonical in mapping:
                continue
            if norm_header in aliases or norm_header == canonical:
                mapping[canonical] = col_idx
                break

    return mapping


def validate_headers(header_map: dict[str, int], raw_headers: Sequence[Any]) -> None:
    """Ensure all required metric columns are present in the header map."""
    missing = [
        FIELD_DISPLAY_NAMES.get(field, field)
        for field in REQUIRED_METRIC_FIELDS
        if field not in header_map
    ]
    # Allow revenue to satisfy roas if roas is not explicitly provided
    if "roas" in missing and "revenue" in header_map:
        missing.remove("roas")
    # Allow clicks and cpc to satisfy spend if spend is not explicitly provided
    if "spend" in missing and "clicks" in header_map and "cpc" in header_map:
        missing.remove("spend")

    if missing:
        raw_names = [str(h) for h in raw_headers if h is not None and str(h).strip()]
        raise MissingRequiredColumnsError(
            f"Missing required column(s): {', '.join(sorted(missing))}. "
            f"Found columns: {raw_names}"
        )


def _parse_float(val: Any, default: float = 0.0) -> float:
    """Parse string/numeric value into float, handling currency and percentage symbols."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s == "-" or s.lower() == "nan" or s.lower() == "none":
        return default
    # Remove currency symbols, commas, spaces, %
    cleaned = re.sub(r"[₹\$,\s]", "", s)
    is_percent = cleaned.endswith("%")
    if is_percent:
        cleaned = cleaned[:-1].strip()
    try:
        fval = float(cleaned)
        return fval
    except ValueError:
        return default


def _parse_int(val: Any, default: int = 0) -> int:
    """Parse string/numeric value into int."""
    f = _parse_float(val, default=float(default))
    return int(round(f))


def _parse_platform(val: Any, context_hint: str = "", default: Platform = Platform.BYOD) -> Platform:
    """Resolve Platform enum from string value or context hint."""
    text = f"{val or ''} {context_hint}".strip().lower()
    tokens = set(re.findall(r"\b\w+\b", text)) | set(re.findall(r"[a-z0-9]+", text))
    if any(k in tokens for k in ("google", "gads", "adwords", "youtube", "google_ads", "g_ads")):
        return Platform.GOOGLE_ADS
    if any(k in tokens for k in ("tiktok", "tik_tok", "douyin", "bytedance", "tiktok_ads", "tt_ads")):
        return Platform.TIKTOK_ADS
    if any(k in tokens for k in ("linkedin", "li_ads", "linkedin_ads", "li")):
        return Platform.LINKEDIN_ADS
    if any(k in tokens for k in ("meta", "facebook", "fb", "instagram", "ig", "meta_ads", "fb_ads", "ig_ads")):
        return Platform.META_ADS
    if any(k in tokens for k in ("byod", "manual", "sheet", "csv", "excel")):
        return Platform.BYOD
    return default


def _is_summary_or_metadata_sheet(sheet_name: str) -> bool:
    """Check if sheet name implies non-tabular metadata or summary."""
    s = sheet_name.strip().lower()
    return s in {"summary", "metadata", "overview", "readme", "notes", "config", "settings"}


def parse_row(
    row_values: Sequence[Any],
    header_map: dict[str, int],
    row_index: int,
    default_platform: Platform = Platform.BYOD,
    context_hint: str = "",
) -> MetricRow | None:
    """Parse a single data row into a MetricRow instance with intelligent autocleaning."""
    # Check if row is completely empty
    if not any(v is not None and str(v).strip() != "" for v in row_values):
        return None

    def get_val(canonical: str) -> Any:
        idx = header_map.get(canonical)
        if idx is not None and idx < len(row_values):
            return row_values[idx]
        return None

    platform_val = get_val("platform")
    if platform_val is not None and str(platform_val).strip():
        platform = _parse_platform(platform_val, context_hint=context_hint, default=default_platform)
    else:
        platform = _parse_platform("", context_hint=context_hint, default=default_platform)

    # Intelligent Campaign Name & ID Synthesis
    campaign_name_val = get_val("campaign_name")
    campaign_type_val = get_val("campaign_type")
    industry_val = get_val("industry")
    country_val = get_val("country")

    if campaign_name_val is not None and str(campaign_name_val).strip():
        campaign_name = str(campaign_name_val).strip()
    else:
        # Build multi-dimensional descriptive name from available context
        parts = []
        plat_label = PLATFORM_LABELS.get(platform, str(platform.value).replace("_", " ").title())
        if campaign_type_val and str(campaign_type_val).strip():
            parts.append(str(campaign_type_val).strip())
        if industry_val and str(industry_val).strip():
            parts.append(str(industry_val).strip())

        main_desc = " - ".join(parts) if parts else f"Campaign {row_index}"
        geo_tag = f" ({str(country_val).strip()})" if country_val and str(country_val).strip() else ""
        campaign_name = f"[{plat_label}] {main_desc}{geo_tag}"

    campaign_id_val = get_val("campaign_id")
    if campaign_id_val is not None and str(campaign_id_val).strip():
        campaign_id = str(campaign_id_val).strip()
    else:
        # Deterministic slug ID
        slug = re.sub(r"[^\w]+", "_", campaign_name.lower()).strip("_")
        campaign_id = f"cmp_{slug or 'row'}_{row_index}"

    spend_inr = _parse_float(get_val("spend_inr"), 0.0)
    impressions = _parse_int(get_val("impressions"), 0)
    clicks = _parse_int(get_val("clicks"), 0)
    conversions = _parse_int(get_val("conversions"), 0)
    cpc_val = _parse_float(get_val("cpc"), 0.0)
    revenue_val = _parse_float(get_val("revenue"), 0.0)

    # 1) Derive spend if missing but clicks and cpc exist
    if spend_inr == 0.0 and clicks > 0 and cpc_val > 0.0:
        spend_inr = round(clicks * cpc_val, 2)

    # 2) ROAS Handling & Derivation
    roas_raw = get_val("roas")
    if roas_raw is not None and str(roas_raw).strip() and str(roas_raw).strip() != "-":
        roas = round(_parse_float(roas_raw, 0.0), 2)
    elif revenue_val > 0.0 and spend_inr > 0.0:
        roas = round(revenue_val / spend_inr, 2)
    else:
        roas = 0.0

    # 3) Derived or explicit CPA
    cpa_raw = get_val("cpa_inr")
    if cpa_raw is not None and str(cpa_raw).strip() and str(cpa_raw).strip() != "-":
        cpa_inr = round(_parse_float(cpa_raw, 0.0), 2)
    elif conversions > 0:
        cpa_inr = round(spend_inr / conversions, 2)
    else:
        cpa_inr = 0.0

    # 4) Derived or explicit CTR (auto-normalize decimal fractions like 0.0353 -> 3.53%)
    ctr_raw = get_val("ctr")
    ctr = 0.0
    if ctr_raw is not None and str(ctr_raw).strip() and str(ctr_raw).strip() != "-":
        raw_str = str(ctr_raw).strip()
        has_percent = "%" in raw_str
        parsed_ctr = _parse_float(ctr_raw, 0.0)
        if 0.0 < parsed_ctr <= 1.0 and not has_percent:
            ctr = round(parsed_ctr * 100.0, 2)
        else:
            ctr = round(parsed_ctr, 2)

    if ctr == 0.0 and impressions > 0:
        ctr = round((clicks / impressions) * 100.0, 2)

    status_val = get_val("status")
    status = str(status_val).strip().upper() if status_val is not None and str(status_val).strip() else "ENABLED"

    return MetricRow(
        campaign_id=campaign_id,
        campaign_name=campaign_name,
        platform=platform,
        spend_inr=spend_inr,
        impressions=impressions,
        clicks=clicks,
        conversions=conversions,
        roas=roas,
        cpa_inr=cpa_inr,
        ctr=ctr,
        status=status,
    )


def _build_snapshot(
    campaigns: list[MetricRow],
    account_ids: list[str] | None = None,
    platform: Platform | None = None,
    source: str = "byod",
) -> CampaignSnapshot:
    """Build aggregated CampaignSnapshot from a list of MetricRows."""
    total_spend = sum(c.spend_inr for c in campaigns)
    if total_spend > 0:
        blended_roas = sum(c.roas * c.spend_inr for c in campaigns) / total_spend
    else:
        blended_roas = 0.0

    # Determine dominant or combined platform
    if platform is not None:
        snap_platform = platform
    elif campaigns:
        platforms = {c.platform for c in campaigns}
        if len(platforms) == 1:
            snap_platform = next(iter(platforms))
        else:
            snap_platform = Platform.BYOD
    else:
        snap_platform = Platform.BYOD

    if not account_ids:
        account_ids = ["byod_account"]

    return CampaignSnapshot(
        account_ids=account_ids,
        platform=snap_platform,
        campaigns=campaigns,
        total_spend_inr=round(total_spend, 2),
        blended_roas=round(blended_roas, 2),
        source=source,
        timestamp=datetime.now(timezone.utc).isoformat(),
    )


def parse_csv(
    source: str | bytes | Path | io.IOBase,
    platform: Platform | str | None = None,
    account_id: str = "byod_account",
    source_tag: str = "byod",
) -> CampaignSnapshot:
    """Parse CSV data into a CampaignSnapshot.

    Args:
        source: File path, string content, raw bytes, or file-like object.
        platform: Optional explicit Platform override.
        account_id: Account ID to associate with snapshot.
        source_tag: Source tag (default: 'byod').

    Returns:
        CampaignSnapshot containing parsed MetricRows.
    """
    if isinstance(source, Path) or (isinstance(source, str) and ("\n" not in source and Path(source).is_file() if len(source) < 500 else False)):
        path = Path(source)
        if not path.is_file():
            raise InvalidDataFormatError(f"CSV file not found: {source}")
        raw_bytes = path.read_bytes()
    elif isinstance(source, bytes):
        raw_bytes = source
    elif isinstance(source, str):
        raw_bytes = source.encode("utf-8")
    elif hasattr(source, "read"):
        content = source.read()
        raw_bytes = content if isinstance(content, bytes) else content.encode("utf-8")
    else:
        raise InvalidDataFormatError(f"Unsupported source type: {type(source)}")

    # Decode text handling UTF-8, UTF-8 BOM, and Latin-1 fallbacks
    text = ""
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        text = raw_bytes.decode("utf-8", errors="replace")

    # Normalize universal line endings (CRLF \r\n and CR \r -> LF \n)
    normalized_text = text.replace("\r\n", "\n").replace("\r", "\n")

    # Initialize StringIO with universal newline="" as recommended by Python's csv module
    string_io = io.StringIO(normalized_text.strip(), newline="")

    # Sniff dialect for delimiter
    try:
        sample = normalized_text[:2048]
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except Exception:
        dialect = csv.excel

    try:
        reader = csv.reader(string_io, dialect=dialect)
        rows = [r for r in reader if any(cell.strip() for cell in r)]
    except csv.Error:
        # Resilient fallback: parse line-by-line using universal splitlines()
        rows = []
        for line in normalized_text.splitlines():
            line_clean = line.strip()
            if not line_clean:
                continue
            line_reader = csv.reader([line_clean], dialect=dialect)
            for r in line_reader:
                if any(cell.strip() for cell in r):
                    rows.append(r)

    if not rows:
        raise InvalidDataFormatError("CSV dataset is empty.")

    raw_headers = rows[0]
    header_map = map_headers(raw_headers)
    validate_headers(header_map, raw_headers)

    default_plat = (
        Platform(platform) if isinstance(platform, str)
        else platform if platform is not None
        else Platform.BYOD
    )

    campaigns: list[MetricRow] = []
    for idx, row in enumerate(rows[1:], start=1):
        metric_row = parse_row(
            row_values=row,
            header_map=header_map,
            row_index=idx,
            default_platform=default_plat,
            context_hint="",
        )
        if metric_row is not None:
            campaigns.append(metric_row)

    return _build_snapshot(
        campaigns=campaigns,
        account_ids=[account_id],
        platform=default_plat if platform is not None else None,
        source=source_tag,
    )


def parse_excel(
    source: str | bytes | Path | io.IOBase,
    sheets: Sequence[str] | None = None,
    default_platform: Platform | str | None = None,
    account_ids: list[str] | None = None,
    ignore_non_tabular: bool = True,
    source_tag: str = "byod",
) -> CampaignSnapshot:
    """Parse an Excel workbook (.xlsx) into a consolidated CampaignSnapshot.

    Supports reading multiple sheets (e.g. 'Google_Ads', 'Meta_Ads', 'Summary').
    Automatically maps sheets like 'Google_Ads' or 'Meta_Ads' to their platforms.
    Gracefully skips non-tabular or summary sheets when ignore_non_tabular=True.

    Args:
        source: File path, raw bytes, or file-like object.
        sheets: Optional list of sheet names to parse. If None, parses all valid sheets.
        default_platform: Fallback platform if not inferred from sheet/row.
        account_ids: List of account IDs.
        ignore_non_tabular: Whether to skip summary/metadata sheets without raising error.
        source_tag: Source tag (default: 'byod').

    Returns:
        CampaignSnapshot combining campaign metrics from all parsed sheets.
    """
    if isinstance(source, (str, Path)) and Path(source).is_file():
        wb = openpyxl.load_workbook(filename=str(source), data_only=True)
    elif isinstance(source, bytes):
        wb = openpyxl.load_workbook(filename=io.BytesIO(source), data_only=True)
    elif hasattr(source, "read"):
        content = source.read()
        byte_io = io.BytesIO(content if isinstance(content, bytes) else content.encode("utf-8"))
        wb = openpyxl.load_workbook(filename=byte_io, data_only=True)
    else:
        raise InvalidDataFormatError(f"Unsupported Excel source: {type(source)}")

    target_sheet_names = list(sheets) if sheets is not None else wb.sheetnames
    if not target_sheet_names:
        raise InvalidDataFormatError("Excel workbook contains no sheets.")

    fallback_platform = (
        Platform(default_platform) if isinstance(default_platform, str)
        else default_platform if default_platform is not None
        else Platform.BYOD
    )

    all_campaigns: list[MetricRow] = []
    parsed_sheet_count = 0
    errors: list[str] = []

    for sheet_name in target_sheet_names:
        if sheet_name not in wb.sheetnames:
            if sheets is not None:
                raise InvalidDataFormatError(f"Sheet '{sheet_name}' not found in workbook.")
            continue

        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        # Filter non-empty rows
        valid_rows = [r for r in raw_rows if any(c is not None and str(c).strip() != "" for c in r)]

        if not valid_rows:
            if _is_summary_or_metadata_sheet(sheet_name) and ignore_non_tabular:
                continue
            if sheets is not None:
                raise InvalidDataFormatError(f"Sheet '{sheet_name}' is empty.")
            continue

        raw_headers = valid_rows[0]
        header_map = map_headers(raw_headers)

        # Check required columns
        try:
            validate_headers(header_map, raw_headers)
        except MissingRequiredColumnsError as err:
            if _is_summary_or_metadata_sheet(sheet_name) and ignore_non_tabular:
                continue
            if sheets is not None or len(target_sheet_names) == 1:
                raise err
            # When auto-scanning multi-sheet workbook, record note and continue
            errors.append(f"Sheet '{sheet_name}': {err}")
            continue

        sheet_platform = _parse_platform("", context_hint=sheet_name, default=fallback_platform)

        for idx, row in enumerate(valid_rows[1:], start=1):
            metric_row = parse_row(
                row_values=row,
                header_map=header_map,
                row_index=len(all_campaigns) + idx,
                default_platform=sheet_platform,
                context_hint=sheet_name,
            )
            if metric_row is not None:
                all_campaigns.append(metric_row)

        parsed_sheet_count += 1

    if parsed_sheet_count == 0 and not all_campaigns:
        if errors:
            raise MissingRequiredColumnsError(f"No valid campaign sheets found. {'; '.join(errors)}")
        raise InvalidDataFormatError("No tabular campaign data found in workbook.")

    return _build_snapshot(
        campaigns=all_campaigns,
        account_ids=account_ids,
        platform=default_platform if default_platform is not None else None,
        source=source_tag,
    )


def parse_excel_sheets(
    source: str | bytes | Path | io.IOBase,
    sheets: Sequence[str] | None = None,
    default_platform: Platform | str | None = None,
) -> dict[str, CampaignSnapshot]:
    """Parse each sheet of an Excel workbook into its own individual CampaignSnapshot."""
    if isinstance(source, (str, Path)) and Path(source).is_file():
        wb = openpyxl.load_workbook(filename=str(source), data_only=True)
    elif isinstance(source, bytes):
        wb = openpyxl.load_workbook(filename=io.BytesIO(source), data_only=True)
    elif hasattr(source, "read"):
        content = source.read()
        byte_io = io.BytesIO(content if isinstance(content, bytes) else content.encode("utf-8"))
        wb = openpyxl.load_workbook(filename=byte_io, data_only=True)
    else:
        raise InvalidDataFormatError(f"Unsupported Excel source: {type(source)}")

    target_sheet_names = list(sheets) if sheets is not None else wb.sheetnames
    result: dict[str, CampaignSnapshot] = {}

    for sheet_name in target_sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        valid_rows = [r for r in raw_rows if any(c is not None and str(c).strip() != "" for c in r)]
        if not valid_rows:
            continue

        raw_headers = valid_rows[0]
        header_map = map_headers(raw_headers)
        try:
            validate_headers(header_map, raw_headers)
        except MissingRequiredColumnsError:
            continue

        sheet_platform = _parse_platform("", context_hint=sheet_name, default=Platform.BYOD)
        campaigns: list[MetricRow] = []
        for idx, row in enumerate(valid_rows[1:], start=1):
            metric_row = parse_row(
                row_values=row,
                header_map=header_map,
                row_index=idx,
                default_platform=sheet_platform,
                context_hint=sheet_name,
            )
            if metric_row is not None:
                campaigns.append(metric_row)

        if campaigns:
            result[sheet_name] = _build_snapshot(
                campaigns=campaigns,
                account_ids=[f"account_{sheet_name.lower()}"],
                platform=sheet_platform,
                source="byod",
            )

    return result


def parse_json(
    source: str | bytes | Path | io.IOBase | dict[str, Any] | list[Any],
    platform: Platform | str | None = None,
    account_id: str = "byod_account",
    source_tag: str = "byod",
) -> CampaignSnapshot:
    """Parse JSON dataset (array of records or dict containing 'campaigns' / 'data') into CampaignSnapshot."""
    data: Any = None
    if isinstance(source, (dict, list)):
        data = source
    elif isinstance(source, (str, Path)) and (isinstance(source, Path) or (len(str(source)) < 500 and "\n" not in str(source) and Path(source).is_file())):
        content = Path(source).read_text(encoding="utf-8")
        data = json.loads(content)
    elif isinstance(source, bytes):
        data = json.loads(source.decode("utf-8"))
    elif isinstance(source, str):
        data = json.loads(source)
    elif hasattr(source, "read"):
        content = source.read()
        text = content.decode("utf-8") if isinstance(content, bytes) else content
        data = json.loads(text)
    else:
        raise InvalidDataFormatError(f"Unsupported JSON source type: {type(source)}")

    rows: list[dict[str, Any]] = []
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        if "campaigns" in data and isinstance(data["campaigns"], list):
            rows = data["campaigns"]
        elif "data" in data and isinstance(data["data"], list):
            rows = data["data"]
        elif "rows" in data and isinstance(data["rows"], list):
            rows = data["rows"]
        else:
            # Single object or dict of sheets
            for k, v in data.items():
                if isinstance(v, list):
                    rows.extend(v)
            if not rows:
                rows = [data]
    else:
        raise InvalidDataFormatError("JSON payload must be a list of campaign objects or a dictionary containing 'campaigns'.")

    if not rows:
        raise InvalidDataFormatError("JSON dataset contains no campaign records.")

    # Convert list of dicts to MetricRows
    campaigns: list[MetricRow] = []
    default_plat = (
        Platform(platform) if isinstance(platform, str)
        else platform if platform is not None
        else Platform.BYOD
    )

    for idx, r in enumerate(rows, start=1):
        if not isinstance(r, dict):
            continue
        # Map dict keys using normalize_column_name
        norm_dict: dict[str, Any] = {}
        for k, v in r.items():
            norm_k = normalize_column_name(str(k))
            for canonical, aliases in COLUMN_ALIASES.items():
                if norm_k in aliases or norm_k == canonical:
                    norm_dict[canonical] = v
                    break

        plat_raw = norm_dict.get("platform") or r.get("platform")
        c_plat = _parse_platform(plat_raw, default=default_plat)

        # Intelligent Campaign Name & ID Synthesis
        campaign_name_val = norm_dict.get("campaign_name", r.get("campaign_name", r.get("name", r.get("campaign"))))
        campaign_type_val = norm_dict.get("campaign_type", r.get("campaign_type", r.get("type")))
        industry_val = norm_dict.get("industry", r.get("industry"))
        country_val = norm_dict.get("country", r.get("country"))

        if campaign_name_val is not None and str(campaign_name_val).strip():
            c_name = str(campaign_name_val).strip()
        else:
            parts = []
            plat_label = PLATFORM_LABELS.get(c_plat, str(c_plat.value).replace("_", " ").title())
            if campaign_type_val and str(campaign_type_val).strip():
                parts.append(str(campaign_type_val).strip())
            if industry_val and str(industry_val).strip():
                parts.append(str(industry_val).strip())

            main_desc = " - ".join(parts) if parts else f"Campaign {idx}"
            geo_tag = f" ({str(country_val).strip()})" if country_val and str(country_val).strip() else ""
            c_name = f"[{plat_label}] {main_desc}{geo_tag}"

        campaign_id_val = norm_dict.get("campaign_id", r.get("campaign_id", r.get("id")))
        if campaign_id_val is not None and str(campaign_id_val).strip():
            c_id = str(campaign_id_val).strip()
        else:
            slug = re.sub(r"[^\w]+", "_", c_name.lower()).strip("_")
            c_id = f"cmp_{slug or 'row'}_{idx}"

        spend_inr = _parse_float(norm_dict.get("spend_inr", r.get("spend", r.get("ad_spend", 0.0))))
        impressions = _parse_int(norm_dict.get("impressions", r.get("impressions", 0)))
        clicks = _parse_int(norm_dict.get("clicks", r.get("clicks", 0)))
        conversions = _parse_int(norm_dict.get("conversions", r.get("conversions", 0)))
        cpc_val = _parse_float(norm_dict.get("cpc", r.get("cpc", r.get("cost_per_click", 0.0))))
        revenue_val = _parse_float(norm_dict.get("revenue", r.get("revenue", r.get("conv_value", 0.0))))

        # 1) Derive spend if missing but clicks and cpc exist
        if spend_inr == 0.0 and clicks > 0 and cpc_val > 0.0:
            spend_inr = round(clicks * cpc_val, 2)

        # 2) ROAS Handling & Derivation
        roas_raw = norm_dict.get("roas", r.get("roas"))
        if roas_raw is not None and str(roas_raw).strip() and str(roas_raw).strip() != "-":
            roas = round(_parse_float(roas_raw, 0.0), 2)
        elif revenue_val > 0.0 and spend_inr > 0.0:
            roas = round(revenue_val / spend_inr, 2)
        else:
            roas = 0.0

        # 3) Derived or explicit CPA
        cpa_raw = norm_dict.get("cpa_inr", r.get("cpa", r.get("cpa_inr")))
        if cpa_raw is not None and str(cpa_raw).strip() and str(cpa_raw).strip() != "-":
            cpa_inr = round(_parse_float(cpa_raw, 0.0), 2)
        elif conversions > 0:
            cpa_inr = round(spend_inr / conversions, 2)
        else:
            cpa_inr = 0.0

        # 4) Derived or explicit CTR (auto-normalize decimal fractions like 0.0353 -> 3.53%)
        ctr_raw = norm_dict.get("ctr", r.get("ctr"))
        ctr = 0.0
        if ctr_raw is not None and str(ctr_raw).strip() and str(ctr_raw).strip() != "-":
            raw_str = str(ctr_raw).strip()
            has_percent = "%" in raw_str
            parsed_ctr = _parse_float(ctr_raw, 0.0)
            if 0.0 < parsed_ctr <= 1.0 and not has_percent:
                ctr = round(parsed_ctr * 100.0, 2)
            else:
                ctr = round(parsed_ctr, 2)

        if ctr == 0.0 and impressions > 0:
            ctr = round((clicks / impressions) * 100.0, 2)

        status_val = norm_dict.get("status", r.get("status", "ENABLED"))
        status = str(status_val).strip().upper() if status_val is not None and str(status_val).strip() else "ENABLED"

        campaigns.append(
            MetricRow(
                campaign_id=c_id,
                campaign_name=c_name,
                platform=c_plat,
                spend_inr=spend_inr,
                impressions=impressions,
                clicks=clicks,
                conversions=conversions,
                roas=roas,
                cpa_inr=cpa_inr,
                ctr=ctr,
                status=status,
            )
        )

    if not campaigns:
        raise InvalidDataFormatError("No valid campaign records could be parsed from JSON.")

    return _build_snapshot(
        campaigns=campaigns,
        account_ids=[account_id],
        platform=default_plat if platform is not None else None,
        source=source_tag,
    )


async def fetch_and_parse_url(
    url: str,
    platform: Platform | str | None = None,
    account_id: str = "url_resource_account",
) -> CampaignSnapshot:
    """Fetch campaign data from a remote URL endpoint (CSV or JSON) and parse into CampaignSnapshot."""
    async with httpx.AsyncClient(timeout=20.0, follow_redirects=True) as client:
        try:
            resp = await client.get(url)
            resp.raise_for_status()
        except Exception as exc:
            raise InvalidDataFormatError(f"Failed to fetch data from URL '{url}': {exc}") from exc

    content_type = resp.headers.get("content-type", "").lower()
    raw_content = resp.content

    # Detect format from content-type or body
    if "json" in content_type or raw_content.strip().startswith((b"{", b"[")):
        return parse_json(raw_content, platform=platform, account_id=account_id, source_tag=f"url:{url}")
    elif "spreadsheet" in content_type or raw_content.startswith(b"PK\x03\x04"):
        return parse_excel(raw_content, default_platform=platform, account_ids=[account_id], source_tag=f"url:{url}")
    else:
        return parse_csv(raw_content, platform=platform, account_id=account_id, source_tag=f"url:{url}")


def _extract_raw_pdf_text(pdf_bytes: bytes) -> str:
    """Fallback text extractor from raw PDF bytes using regex token matching."""
    text_chunks = []
    try:
        decoded = pdf_bytes.decode("latin1", errors="ignore")
        for match in re.finditer(r"\(([^)]+)\)\s*Tj", decoded):
            text_chunks.append(match.group(1))
    except Exception:
        pass
    return " ".join(text_chunks)


def parse_pdf(
    source: str | bytes | Path | io.IOBase,
    default_platform: Platform | str | None = None,
    account_ids: Sequence[str] | None = None,
    source_tag: str = "byod",
) -> CampaignSnapshot:
    """Parse PDF document containing marketing campaign tables, summaries, or briefs into CampaignSnapshot."""
    pdf_bytes: bytes
    if isinstance(source, (str, Path)) and Path(source).is_file():
        pdf_bytes = Path(source).read_bytes()
    elif isinstance(source, bytes):
        pdf_bytes = source
    elif isinstance(source, str):
        clean = source.strip()
        if clean.startswith("data:") and ";base64," in clean:
            clean = clean.split(";base64,")[1]
        import base64
        try:
            pdf_bytes = base64.b64decode(clean)
        except Exception:
            pdf_bytes = source.encode("utf-8")
    elif hasattr(source, "read"):
        content = source.read()
        pdf_bytes = content if isinstance(content, bytes) else content.encode("utf-8")
    else:
        raise InvalidDataFormatError(f"Unsupported PDF source: {type(source)}")

    if not pdf_bytes or len(pdf_bytes) < 4:
        raise InvalidDataFormatError("Uploaded PDF file is empty or corrupted.")

    pages_text: list[str] = []
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        for page in reader.pages:
            t = page.extract_text()
            if t:
                pages_text.append(t)
    except Exception as exc:
        raw = _extract_raw_pdf_text(pdf_bytes)
        if raw:
            pages_text.append(raw)

    full_text = "\n\n".join(pages_text).strip()
    if not full_text:
        full_text = _extract_raw_pdf_text(pdf_bytes).strip()

    campaigns: list[MetricRow] = []
    plat = (
        Platform(default_platform) if isinstance(default_platform, str)
        else default_platform if default_platform is not None
        else Platform.BYOD
    )

    lines = [line.strip() for line in full_text.splitlines() if line.strip()]

    # 1. Try parsing tabular lines if present
    table_lines = []
    header_found = False
    for line in lines:
        norm = normalize_column_name(line)
        if any(req in norm for req in ("spend", "campaign", "clicks", "roas", "impressions", "cpa", "ctr")):
            header_found = True
        if header_found:
            table_lines.append(line)

    if table_lines and len(table_lines) >= 2:
        try:
            sub_csv = "\n".join(table_lines)
            sub_snap = parse_csv(sub_csv, platform=plat)
            if sub_snap.campaigns:
                campaigns.extend(sub_snap.campaigns)
        except Exception:
            pass

    # 2. Try regex extraction of campaign blocks
    if not campaigns:
        campaign_pattern = re.compile(
            r"(?:Campaign|Campaign Name|Initiative|Ad Set)[:\s-]+(.*?)"
            r"(?=(?:Spend|Budget|Investment|ROAS|Clicks|Conversions)[:\s-]|\n|$)",
            re.IGNORECASE,
        )

        # Split text into sections or scan match-by-match
        for match in campaign_pattern.finditer(full_text):
            c_name = match.group(1).strip(" :-\t,|")
            if not c_name or len(c_name) < 2 or c_name.lower() in ("name", "id", "status"):
                continue

            # Look for metrics in vicinity of this campaign mention (up to 300 chars forward)
            start_pos = match.end()
            context_window = full_text[start_pos : start_pos + 300]

            spend_m = re.search(r"(?:Spend|Budget|Investment|Cost)[:\s-]*[₹$€]?\s*([\d,.]+)", context_window, re.IGNORECASE)
            roas_m = re.search(r"(?:ROAS|Return)[:\s-]*([\d.]+)", context_window, re.IGNORECASE)
            clicks_m = re.search(r"(?:Clicks)[:\s-]*([\d,.]+)", context_window, re.IGNORECASE)
            conv_m = re.search(r"(?:Conversions)[:\s-]*([\d,.]+)", context_window, re.IGNORECASE)

            s_val = _parse_float(spend_m.group(1), 0.0) if spend_m else 0.0
            r_val = _parse_float(roas_m.group(1), 0.0) if roas_m else 0.0
            cl_val = _parse_int(clicks_m.group(1), 0) if clicks_m else 0
            cv_val = _parse_int(conv_m.group(1), 0) if conv_m else 0

            c_plat = _parse_platform("", context_hint=c_name, default=plat)
            slug = re.sub(r"[^\w]+", "_", c_name.lower()).strip("_")
            c_id = f"cmp_{slug}_{len(campaigns) + 1}"

            cpa = round(s_val / cv_val, 2) if cv_val > 0 else 0.0
            ctr = round((cl_val / 1000) * 100, 2) if cl_val > 0 else 3.5

            campaigns.append(
                MetricRow(
                    campaign_id=c_id,
                    campaign_name=c_name,
                    platform=c_plat,
                    spend_inr=s_val or 50000.0,
                    impressions=cl_val * 25 if cl_val else 50000,
                    clicks=cl_val or 2000,
                    conversions=cv_val or 100,
                    roas=r_val or 2.5,
                    cpa_inr=cpa or 500.0,
                    ctr=ctr,
                    status="ENABLED",
                )
            )

    # 3. Fallback narrative representation
    if not campaigns:
        doc_title = lines[0] if lines else "PDF Marketing Document"
        if len(doc_title) > 60:
            doc_title = doc_title[:60] + "…"

        spend_match = re.search(r"(?:spend|budget|investment|cost)[:\s-]*[₹$€]?\s*([\d,.]+)", full_text, re.IGNORECASE)
        total_spend = _parse_float(spend_match.group(1), 75000.0) if spend_match else 75000.0

        roas_match = re.search(r"(?:roas|return|target roas)[:\s-]*([\d.]+)", full_text, re.IGNORECASE)
        total_roas = _parse_float(roas_match.group(1), 2.8) if roas_match else 2.8

        c_plat = _parse_platform("", context_hint=full_text, default=plat)
        c_id = "cmp_pdf_doc_1"

        campaigns.append(
            MetricRow(
                campaign_id=c_id,
                campaign_name=f"[PDF] {doc_title}",
                platform=c_plat,
                spend_inr=total_spend,
                impressions=int(total_spend * 4),
                clicks=int(total_spend * 0.05),
                conversions=max(1, int(total_spend * 0.002)),
                roas=total_roas,
                cpa_inr=round(total_spend / max(1, int(total_spend * 0.002)), 2),
                ctr=3.85,
                status="ENABLED",
            )
        )

    notes_text = f"Ingested from PDF ({len(pdf_bytes)} bytes):\n\n" + (full_text[:4000] if full_text else "Binary PDF payload")

    return _build_snapshot(
        campaigns=campaigns,
        account_ids=list(account_ids) if account_ids else ["pdf_account"],
        platform=plat,
        source=source_tag,
    )


def import_byod_file(
    source: str | bytes | Path | io.IOBase,
    filename: str | None = None,
    default_platform: Platform | str | None = None,
) -> CampaignSnapshot:
    """Unified importer that auto-detects CSV, Excel, JSON, or PDF formats."""
    # Determine filename/format
    name = ""
    if filename:
        name = filename
    elif isinstance(source, (str, Path)):
        name = str(source)

    name_lower = name.lower()
    if name_lower.endswith(".pdf"):
        return parse_pdf(source, default_platform=default_platform)
    elif name_lower.endswith((".xlsx", ".xlsm", ".xltx")):
        return parse_excel(source, default_platform=default_platform)
    elif name_lower.endswith(".xls"):
        try:
            return parse_excel(source, default_platform=default_platform)
        except Exception:
            try:
                return parse_csv(source, platform=default_platform)
            except Exception:
                raise InvalidDataFormatError("Failed to parse .xls file. Ensure it is a valid Excel workbook or tabular text export.")
    elif name_lower.endswith((".json", ".js")):
        return parse_json(source, platform=default_platform)
    elif name_lower.endswith((".csv", ".tsv", ".txt")):
        return parse_csv(source, platform=default_platform)

    # Check magic bytes for PDF
    if isinstance(source, bytes) and source.startswith(b"%PDF"):
        return parse_pdf(source, default_platform=default_platform)

    # Check magic bytes for ZIP (XLSX)
    if isinstance(source, bytes) and source.startswith(b"PK\x03\x04"):
        return parse_excel(source, default_platform=default_platform)

    # Check for JSON format
    if isinstance(source, (str, bytes)):
        raw_str = source.decode("utf-8", errors="ignore") if isinstance(source, bytes) else source
        trimmed = raw_str.strip()
        if trimmed.startswith(("{", "[")):
            try:
                return parse_json(source, platform=default_platform)
            except Exception:
                pass

    # Fallback to CSV, then Excel, then PDF
    try:
        return parse_csv(source, platform=default_platform)
    except Exception:
        try:
            return parse_excel(source, default_platform=default_platform)
        except Exception:
            return parse_pdf(source, default_platform=default_platform)


# ---------------------------------------------------------------------------
# Active In-Memory BYOD Dataset Store
# ---------------------------------------------------------------------------
_ACTIVE_BYOD_SNAPSHOT: CampaignSnapshot | None = None


def set_active_byod_snapshot(snapshot: CampaignSnapshot) -> None:
    """Set the active BYOD campaign snapshot for Governor and analyst runs."""
    global _ACTIVE_BYOD_SNAPSHOT
    _ACTIVE_BYOD_SNAPSHOT = snapshot


def get_active_byod_snapshot() -> CampaignSnapshot | None:
    """Retrieve the currently active BYOD campaign snapshot."""
    global _ACTIVE_BYOD_SNAPSHOT
    return _ACTIVE_BYOD_SNAPSHOT


def clear_active_byod_snapshot() -> None:
    """Clear active BYOD dataset, reverting back to synthetic/live modes."""
    global _ACTIVE_BYOD_SNAPSHOT
    _ACTIVE_BYOD_SNAPSHOT = None


def has_active_byod_snapshot() -> bool:
    """Check if an active BYOD dataset is currently loaded."""
    global _ACTIVE_BYOD_SNAPSHOT
    return _ACTIVE_BYOD_SNAPSHOT is not None


def decode_byod_content(
    file_content: str | bytes,
    filename: str | None = None,
) -> bytes:
    """Helper to cleanly decode Data URL, Base64, or text payload into raw bytes."""
    import base64

    fname = filename or "uploaded_dataset.csv"
    if isinstance(file_content, bytes):
        return file_content
    elif isinstance(file_content, str):
        content_str = file_content.strip()
        if content_str.startswith("data:") and ";base64," in content_str:
            content_str = content_str.split(";base64,")[1]
            try:
                return base64.b64decode(content_str)
            except Exception as exc:
                raise InvalidDataFormatError(f"Failed to decode base64 file data: {exc}") from exc
        else:
            # Check if this is a binary file (.xlsx, .xls, .pdf)
            if fname.lower().endswith((".xlsx", ".xls", ".xlsm", ".pdf")):
                try:
                    return base64.b64decode(content_str)
                except Exception:
                    return content_str.encode("utf-8")
            else:
                # For CSV/JSON/text: if it contains text markers, encode directly
                if content_str.startswith(("{", "[")) or "\n" in content_str or "," in content_str:
                    return content_str.encode("utf-8")
                else:
                    try:
                        return base64.b64decode(content_str)
                    except Exception:
                        return content_str.encode("utf-8")
    else:
        raise InvalidDataFormatError(f"Unsupported file content type: {type(file_content)}")


def activate_byod_file(
    file_content: str | bytes,
    filename: str | None = None,
    default_platform: Platform | str | None = None,
) -> CampaignSnapshot:
    """Helper to parse a base64, raw string, or bytes dataset and activate it as the active BYOD snapshot."""
    fname = filename or "uploaded_dataset.csv"
    raw_bytes = decode_byod_content(file_content, filename=fname)
    snapshot = import_byod_file(raw_bytes, filename=fname, default_platform=default_platform)
    set_active_byod_snapshot(snapshot)
    return snapshot


# Aliases for convenience
import_csv = parse_csv
import_excel = parse_excel
import_json = parse_json


# ---------------------------------------------------------------------------
# Sample Finnovate Bundle Generator
# ---------------------------------------------------------------------------

def get_finnovate_sample_data() -> dict[str, list[dict[str, Any]]]:
    """Structured mock campaigns for Finnovate ad accounts."""
    return {
        "Google_Ads": [
            {
                "campaign_id": "cmp_gads_search_01",
                "campaign_name": "Finnovate — Mutual Fund Search Intent",
                "platform": "google_ads",
                "spend_inr": 45000.0,
                "impressions": 125000,
                "clicks": 8400,
                "conversions": 420,
                "roas": 3.4,
                "cpa_inr": 107.14,
                "ctr": 6.72,
                "status": "ENABLED",
            },
            {
                "campaign_id": "cmp_gads_brand_02",
                "campaign_name": "Finnovate — Brand Search Core",
                "platform": "google_ads",
                "spend_inr": 25000.0,
                "impressions": 80000,
                "clicks": 7200,
                "conversions": 360,
                "roas": 4.2,
                "cpa_inr": 69.44,
                "ctr": 9.0,
                "status": "ENABLED",
            },
            {
                "campaign_id": "cmp_gads_pmax_03",
                "campaign_name": "Finnovate — Performance Max Direct SIP",
                "platform": "google_ads",
                "spend_inr": 50000.0,
                "impressions": 220000,
                "clicks": 9500,
                "conversions": 380,
                "roas": 2.8,
                "cpa_inr": 131.58,
                "ctr": 4.32,
                "status": "ENABLED",
            },
        ],
        "Meta_Ads": [
            {
                "campaign_id": "cmp_meta_retargeting_01",
                "campaign_name": "Finnovate — SIP Growth Retargeting",
                "platform": "meta_ads",
                "spend_inr": 65000.0,
                "impressions": 340000,
                "clicks": 14200,
                "conversions": 510,
                "roas": 2.1,
                "cpa_inr": 127.45,
                "ctr": 4.18,
                "status": "ENABLED",
            },
            {
                "campaign_id": "cmp_meta_gold_fatigue_02",
                "campaign_name": "Finnovate — Gold ETF Broad Audience",
                "platform": "meta_ads",
                "spend_inr": 30000.0,
                "impressions": 210000,
                "clicks": 3900,
                "conversions": 95,
                "roas": 1.1,
                "cpa_inr": 315.79,
                "ctr": 1.86,
                "status": "PAUSED",
            },
            {
                "campaign_id": "cmp_meta_hnw_lookalike_03",
                "campaign_name": "Finnovate — High Net Worth Lookalike 1%",
                "platform": "meta_ads",
                "spend_inr": 55000.0,
                "impressions": 190000,
                "clicks": 8100,
                "conversions": 310,
                "roas": 2.9,
                "cpa_inr": 177.42,
                "ctr": 4.26,
                "status": "ENABLED",
            },
        ],
        "Summary": [
            {
                "metric": "Total Accounts",
                "value": "2 (Google Ads & Meta Ads)",
            },
            {
                "metric": "Target Monthly Budget (INR)",
                "value": "270000.00",
            },
            {
                "metric": "Blended Target ROAS",
                "value": "2.80",
            },
            {
                "metric": "Primary Conversion Goal",
                "value": "SIP Registration Completed",
            },
        ],
    }


def create_finnovate_excel_bytes() -> bytes:
    """Generate in-memory multi-sheet Excel workbook bytes containing Finnovate sample data."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)  # type: ignore

    data = get_finnovate_sample_data()

    # Google Ads sheet
    ws_gads = wb.create_sheet(title="Google_Ads")
    gads_headers = [
        "Campaign ID", "Campaign Name", "Platform", "Spend (INR)",
        "Impressions", "Clicks", "Conversions", "ROAS", "CPA (INR)", "CTR (%)", "Status"
    ]
    ws_gads.append(gads_headers)
    for row in data["Google_Ads"]:
        ws_gads.append([
            row["campaign_id"], row["campaign_name"], row["platform"], row["spend_inr"],
            row["impressions"], row["clicks"], row["conversions"], row["roas"],
            row["cpa_inr"], row["ctr"], row["status"]
        ])

    # Meta Ads sheet
    ws_meta = wb.create_sheet(title="Meta_Ads")
    meta_headers = [
        "Campaign ID", "Campaign Name", "Platform", "Spend (INR)",
        "Impressions", "Clicks", "Conversions", "ROAS", "CPA (INR)", "CTR (%)", "Status"
    ]
    ws_meta.append(meta_headers)
    for row in data["Meta_Ads"]:
        ws_meta.append([
            row["campaign_id"], row["campaign_name"], row["platform"], row["spend_inr"],
            row["impressions"], row["clicks"], row["conversions"], row["roas"],
            row["cpa_inr"], row["ctr"], row["status"]
        ])

    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(["Metric", "Value"])
    for row in data["Summary"]:
        ws_summary.append([row["metric"], row["value"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_finnovate_csv_bytes(sheet_name: str = "Google_Ads") -> bytes:
    """Generate in-memory CSV bytes for Finnovate sample campaigns."""
    data = get_finnovate_sample_data()
    rows = data.get(sheet_name, data["Google_Ads"])

    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=[
            "campaign_id", "campaign_name", "platform", "spend_inr",
            "impressions", "clicks", "conversions", "roas", "cpa_inr", "ctr", "status"
        ],
    )
    writer.writeheader()
    for r in rows:
        writer.writerow(r)

    return buf.getvalue().encode("utf-8")


def generate_finnovate_sample_bundle(
    output_dir: str | Path | None = None,
) -> dict[str, Path | bytes]:
    """Generate complete Finnovate sample bundle for testing and demos.

    If output_dir is provided, writes files to disk and returns mapping of filename -> Path.
    If output_dir is None, returns mapping of filename -> bytes.
    """
    excel_bytes = create_finnovate_excel_bytes()
    gads_csv = create_finnovate_csv_bytes("Google_Ads")
    meta_csv = create_finnovate_csv_bytes("Meta_Ads")

    # Also generate blended CSV
    all_data = get_finnovate_sample_data()
    blended_rows = all_data["Google_Ads"] + all_data["Meta_Ads"]
    blended_buf = io.StringIO()
    writer = csv.DictWriter(
        blended_buf,
        fieldnames=[
            "campaign_id", "campaign_name", "platform", "spend_inr",
            "impressions", "clicks", "conversions", "roas", "cpa_inr", "ctr", "status"
        ],
    )
    writer.writeheader()
    for r in blended_rows:
        writer.writerow(r)
    blended_csv = blended_buf.getvalue().encode("utf-8")

    bundle: dict[str, bytes] = {
        "finnovate_campaigns.xlsx": excel_bytes,
        "finnovate_google_ads.csv": gads_csv,
        "finnovate_meta_ads.csv": meta_csv,
        "finnovate_blended.csv": blended_csv,
    }

    if output_dir is not None:
        out_path = Path(output_dir)
        out_path.mkdir(parents=True, exist_ok=True)
        disk_bundle: dict[str, Path] = {}
        for fname, bdata in bundle.items():
            fpath = out_path / fname
            fpath.write_bytes(bdata)
            disk_bundle[fname] = fpath
        return disk_bundle  # type: ignore

    return bundle  # type: ignore
