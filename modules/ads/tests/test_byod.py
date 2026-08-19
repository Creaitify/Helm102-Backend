"""Comprehensive unit tests for BYOD campaign importer and GAQL parser."""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest

from modules.ads.byod_importer import (
    InvalidDataFormatError,
    MissingRequiredColumnsError,
    generate_finnovate_sample_bundle,
    get_finnovate_sample_data,
    import_byod_file,
    import_csv,
    import_excel,
    parse_csv,
    parse_excel,
    parse_excel_sheets,
)
from modules.ads.contracts import Platform
from modules.ads.gaql import (
    generate_campaign_performance_gaql,
    parse_gaql_response,
)


# ---------------------------------------------------------------------------
# CSV Parsing Tests
# ---------------------------------------------------------------------------

def test_parse_csv_basic():
    csv_content = """campaign_id,campaign_name,platform,spend_inr,impressions,clicks,conversions,roas,cpa_inr,ctr,status
cmp_01,Search Intent Campaign,google_ads,45000,125000,8400,420,3.4,107.14,6.72,ENABLED
cmp_02,Retargeting Campaign,meta_ads,65000,340000,14200,510,2.1,127.45,4.18,ENABLED
"""
    snapshot = parse_csv(csv_content)
    assert len(snapshot.campaigns) == 2
    assert snapshot.total_spend_inr == 110000.0
    # Blended ROAS = (45000*3.4 + 65000*2.1) / 110000 = (153000 + 136500) / 110000 = 289500 / 110000 = 2.6318 -> 2.63
    assert snapshot.blended_roas == 2.63
    assert snapshot.campaigns[0].campaign_id == "cmp_01"
    assert snapshot.campaigns[0].platform == Platform.GOOGLE_ADS
    assert snapshot.campaigns[0].spend_inr == 45000.0
    assert snapshot.campaigns[0].roas == 3.4
    assert snapshot.campaigns[1].campaign_id == "cmp_02"
    assert snapshot.campaigns[1].platform == Platform.META_ADS


def test_parse_csv_with_aliases_and_currency_formatting():
    """Test handling of headers with aliases and values with currency signs, commas, and percentage."""
    csv_content = """Campaign Name,Cost (INR),Views,Link Clicks,Results,Conv. Value / Cost,Status
Mutual Funds High Intent,"₹ 50,000.00","200,000","10,000","500","3.50",Active
Gold ETF Growth,"₹ 30,000.00","150,000","6,000","150","1.50",Paused
"""
    snapshot = parse_csv(csv_content, platform=Platform.GOOGLE_ADS)
    assert len(snapshot.campaigns) == 2
    assert snapshot.campaigns[0].spend_inr == 50000.0
    assert snapshot.campaigns[0].impressions == 200000
    assert snapshot.campaigns[0].clicks == 10000
    assert snapshot.campaigns[0].conversions == 500
    assert snapshot.campaigns[0].roas == 3.5
    assert snapshot.campaigns[0].status == "ACTIVE"
    # Auto-calculated CPA = 50000 / 500 = 100.0
    assert snapshot.campaigns[0].cpa_inr == 100.0
    # Auto-calculated CTR = (10000 / 200000) * 100 = 5.0
    assert snapshot.campaigns[0].ctr == 5.0


def test_parse_csv_auto_cpa_ctr_calculation():
    """Verify CPA and CTR are accurately computed when omitted in source data."""
    csv_content = """campaign_name,spend,clicks,conversions,roas,impressions
SIP Growth,20000,1000,50,2.5,20000
"""
    snapshot = parse_csv(csv_content)
    camp = snapshot.campaigns[0]
    assert camp.cpa_inr == 400.0  # 20000 / 50
    assert camp.ctr == 5.0  # (1000 / 20000) * 100


def test_parse_csv_zero_conversions_impressions_no_division_error():
    """Verify safe zero division when conversions or impressions are 0."""
    csv_content = """campaign_name,spend,clicks,conversions,roas,impressions
Brand Zero,5000,0,0,0.0,0
"""
    snapshot = parse_csv(csv_content)
    camp = snapshot.campaigns[0]
    assert camp.cpa_inr == 0.0
    assert camp.ctr == 0.0


# ---------------------------------------------------------------------------
# Validation Tests for Required Columns
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "missing_col,header_str",
    [
        ("spend", "campaign_name,clicks,conversions,roas"),
        ("roas", "campaign_name,spend,clicks,conversions"),
        ("clicks", "campaign_name,spend,conversions,roas"),
        ("conversions", "campaign_name,spend,clicks,roas"),
    ],
)
def test_missing_required_column_raises_error(missing_col: str, header_str: str):
    csv_content = f"{header_str}\nCampA,1000,50,2.0\n"
    with pytest.raises(MissingRequiredColumnsError) as exc_info:
        parse_csv(csv_content)
    assert missing_col in str(exc_info.value).lower()


def test_missing_multiple_required_columns_reports_all():
    csv_content = "campaign_name,impressions\nCampA,10000\n"
    with pytest.raises(MissingRequiredColumnsError) as exc_info:
        parse_csv(csv_content)
    msg = str(exc_info.value).lower()
    assert "spend" in msg
    assert "roas" in msg
    assert "clicks" in msg
    assert "conversions" in msg


# ---------------------------------------------------------------------------
# Excel (.xlsx) Multi-Sheet Tests
# ---------------------------------------------------------------------------

def test_parse_excel_multi_sheet(tmp_path: Path):
    """Test multi-sheet workbook reading Google_Ads, Meta_Ads, and skipping Summary."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore

    # Sheet 1: Google_Ads
    ws_gads = wb.create_sheet(title="Google_Ads")
    ws_gads.append(["Campaign ID", "Campaign Name", "Spend (INR)", "Impressions", "Clicks", "Conversions", "ROAS"])
    ws_gads.append(["g_01", "Google Search Fund", 45000, 125000, 8400, 420, 3.4])
    ws_gads.append(["g_02", "Google Brand", 25000, 80000, 7200, 360, 4.2])

    # Sheet 2: Meta_Ads
    ws_meta = wb.create_sheet(title="Meta_Ads")
    ws_meta.append(["Campaign ID", "Campaign Name", "Spend (INR)", "Impressions", "Clicks", "Conversions", "ROAS"])
    ws_meta.append(["m_01", "Meta SIP Retargeting", 65000, 340000, 14200, 510, 2.1])

    # Sheet 3: Summary (metadata / non-tabular)
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.append(["Key", "Value"])
    ws_sum.append(["Owner", "AdOps Team"])
    ws_sum.append(["Quarter", "Q3"])

    file_path = tmp_path / "campaigns.xlsx"
    wb.save(str(file_path))

    snapshot = parse_excel(file_path)
    assert len(snapshot.campaigns) == 3
    assert snapshot.total_spend_inr == 135000.0

    gads_camps = [c for c in snapshot.campaigns if c.platform == Platform.GOOGLE_ADS]
    meta_camps = [c for c in snapshot.campaigns if c.platform == Platform.META_ADS]
    assert len(gads_camps) == 2
    assert len(meta_camps) == 1
    assert gads_camps[0].campaign_id == "g_01"
    assert meta_camps[0].campaign_id == "m_01"


def test_parse_excel_sheets_dict(tmp_path: Path):
    """Test parse_excel_sheets returns distinct CampaignSnapshots per valid sheet."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore

    ws_gads = wb.create_sheet(title="Google_Ads")
    ws_gads.append(["Campaign Name", "Spend", "Clicks", "Conversions", "ROAS"])
    ws_gads.append(["G Search", 10000, 500, 25, 3.0])

    ws_meta = wb.create_sheet(title="Meta_Ads")
    ws_meta.append(["Campaign Name", "Spend", "Clicks", "Conversions", "ROAS"])
    ws_meta.append(["M Retargeting", 15000, 800, 30, 2.0])

    file_path = tmp_path / "split_sheets.xlsx"
    wb.save(str(file_path))

    sheets_dict = parse_excel_sheets(file_path)
    assert "Google_Ads" in sheets_dict
    assert "Meta_Ads" in sheets_dict
    assert len(sheets_dict["Google_Ads"].campaigns) == 1
    assert len(sheets_dict["Meta_Ads"].campaigns) == 1
    assert sheets_dict["Google_Ads"].total_spend_inr == 10000.0
    assert sheets_dict["Meta_Ads"].total_spend_inr == 15000.0


def test_parse_excel_missing_columns_in_single_sheet(tmp_path: Path):
    """Ensure MissingRequiredColumnsError is raised for invalid single sheet Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campaigns"  # type: ignore
    ws.append(["Campaign Name", "Spend", "Clicks"])  # Missing roas & conversions  # type: ignore
    ws.append(["Camp1", 1000, 50])  # type: ignore

    file_path = tmp_path / "invalid.xlsx"
    wb.save(str(file_path))

    with pytest.raises(MissingRequiredColumnsError) as exc_info:
        parse_excel(file_path)
    msg = str(exc_info.value).lower()
    assert "conversions" in msg
    assert "roas" in msg


# ---------------------------------------------------------------------------
# Finnovate Bundle Generator & Importer Integration Tests
# ---------------------------------------------------------------------------

def test_generate_finnovate_sample_bundle_in_memory():
    """Verify in-memory bundle generation creates valid binary artifacts."""
    bundle = generate_finnovate_sample_bundle()
    assert isinstance(bundle, dict)
    assert "finnovate_campaigns.xlsx" in bundle
    assert "finnovate_google_ads.csv" in bundle
    assert "finnovate_meta_ads.csv" in bundle
    assert "finnovate_blended.csv" in bundle

    # Test importing each artifact back
    snap_xlsx = import_byod_file(bundle["finnovate_campaigns.xlsx"], filename="finnovate_campaigns.xlsx")
    assert len(snap_xlsx.campaigns) == 6
    assert snap_xlsx.total_spend_inr == 270000.0

    snap_gads_csv = import_byod_file(bundle["finnovate_google_ads.csv"], filename="finnovate_google_ads.csv")
    assert len(snap_gads_csv.campaigns) == 3
    assert snap_gads_csv.total_spend_inr == 120000.0

    snap_meta_csv = import_byod_file(bundle["finnovate_meta_ads.csv"], filename="finnovate_meta_ads.csv")
    assert len(snap_meta_csv.campaigns) == 3
    assert snap_meta_csv.total_spend_inr == 150000.0

    snap_blended_csv = import_byod_file(bundle["finnovate_blended.csv"], filename="finnovate_blended.csv")
    assert len(snap_blended_csv.campaigns) == 6
    assert snap_blended_csv.total_spend_inr == 270000.0


def test_generate_finnovate_sample_bundle_to_disk(tmp_path: Path):
    """Verify bundle writes cleanly to disk and parses properly."""
    disk_bundle = generate_finnovate_sample_bundle(output_dir=tmp_path)
    assert isinstance(disk_bundle, dict)

    xlsx_path = disk_bundle["finnovate_campaigns.xlsx"]
    assert isinstance(xlsx_path, Path)
    assert xlsx_path.is_file()

    snap = import_excel(xlsx_path)
    assert len(snap.campaigns) == 6
    assert snap.total_spend_inr == 270000.0
    assert snap.source == "byod"


def test_import_byod_file_format_detection(tmp_path: Path):
    """Test auto-format detection with paths, filenames, and byte contents."""
    bundle = generate_finnovate_sample_bundle()

    # XLSX from bytes without explicit filename (using magic header)
    snap_bytes = import_byod_file(bundle["finnovate_campaigns.xlsx"])
    assert len(snap_bytes.campaigns) == 6

    # CSV from bytes
    snap_csv_bytes = import_byod_file(bundle["finnovate_google_ads.csv"], filename="data.csv")
    assert len(snap_csv_bytes.campaigns) == 3


def test_empty_or_corrupted_inputs():
    """Verify clean exceptions for empty or malformed datasets."""
    with pytest.raises(InvalidDataFormatError):
        parse_csv("")

    with pytest.raises(InvalidDataFormatError):
        parse_csv("   \n\n   ")

    with pytest.raises(InvalidDataFormatError):
        parse_csv(Path("c:/non_existent_path_xyz123.csv"))


# ---------------------------------------------------------------------------
# GAQL Query Generator & Response Parser Tests
# ---------------------------------------------------------------------------

def test_generate_campaign_performance_gaql():
    query_default = generate_campaign_performance_gaql()
    assert "SELECT campaign.id" in query_default
    assert "segments.date DURING LAST_30_DAYS" in query_default
    assert "campaign.status = 'ENABLED'" in query_default

    query_custom = generate_campaign_performance_gaql(
        date_range="LAST_7_DAYS",
        campaign_ids=["12345", "67890"],
        status_filter="PAUSED",
    )
    assert "segments.date DURING LAST_7_DAYS" in query_custom
    assert "campaign.id IN (12345, 67890)" in query_custom
    assert "campaign.status = 'PAUSED'" in query_custom


def test_parse_gaql_response():
    mock_rows = [
        {
            "campaign": {"id": 1001, "name": "Search Performance", "status": "ENABLED"},
            "metrics": {
                "cost_micros": 50000000000,  # 50,000 INR
                "impressions": 100000,
                "clicks": 5000,
                "conversions": 250,
                "conversions_value": 150000.0,
                "ctr": 0.05,  # 5.0%
                "cost_per_conversion": 200000000,  # 200 INR
            },
        },
        {
            "campaign": {"id": 1002, "name": "Brand Core", "status": "ENABLED"},
            "metrics": {
                "cost_micros": 20000000000,  # 20,000 INR
                "impressions": 50000,
                "clicks": 4000,
                "conversions": 200,
                "conversions_value": 80000.0,
                "ctr": 0.08,  # 8.0%
                "cost_per_conversion": 100000000,  # 100 INR
            },
        },
    ]

    snapshot = parse_gaql_response(mock_rows, account_id="cust_finnovate_123")
    assert len(snapshot.campaigns) == 2
    assert snapshot.total_spend_inr == 70000.0
    assert snapshot.campaigns[0].campaign_id == "1001"
    assert snapshot.campaigns[0].spend_inr == 50000.0
    assert snapshot.campaigns[0].roas == 3.0  # 150000 / 50000
    assert snapshot.campaigns[0].cpa_inr == 200.0
    assert snapshot.campaigns[0].ctr == 5.0
    assert snapshot.campaigns[1].campaign_id == "1002"
    assert snapshot.campaigns[1].spend_inr == 20000.0
    assert snapshot.campaigns[1].roas == 4.0  # 80000 / 20000
    assert snapshot.campaigns[1].cpa_inr == 100.0
    assert snapshot.campaigns[1].ctr == 8.0


# ---------------------------------------------------------------------------
# Metric Auto-Derivations & Multi-Platform Synthesis Unit Tests
# ---------------------------------------------------------------------------

def test_parse_csv_spend_auto_derivation_from_clicks_and_cpc():
    """Verify spend is auto-derived when spend column is omitted but clicks and cpc are present."""
    csv_content = """campaign_name,clicks,cpc,conversions,roas,impressions
Search Intent,500,2.50,25,3.0,10000
"""
    snapshot = parse_csv(csv_content)
    assert len(snapshot.campaigns) == 1
    camp = snapshot.campaigns[0]
    # spend_inr = round(500 * 2.50, 2) = 1250.0
    assert camp.spend_inr == 1250.0
    assert camp.roas == 3.0
    # cpa = 1250 / 25 = 50.0
    assert camp.cpa_inr == 50.0
    # ctr = (500 / 10000) * 100 = 5.0
    assert camp.ctr == 5.0


def test_parse_csv_chained_metric_derivation():
    """Verify chained derivation: spend derived from clicks*cpc, then roas derived from revenue/spend."""
    csv_content = """campaign_name,clicks,cpc,conversions,revenue,impressions
Growth Scale,800,1.75,40,4200.0,20000
"""
    snapshot = parse_csv(csv_content)
    assert len(snapshot.campaigns) == 1
    camp = snapshot.campaigns[0]
    # spend = round(800 * 1.75, 2) = 1400.0
    assert camp.spend_inr == 1400.0
    # roas = round(4200.0 / 1400.0, 2) = 3.0
    assert camp.roas == 3.0
    # cpa = round(1400.0 / 40, 2) = 35.0
    assert camp.cpa_inr == 35.0


def test_parse_csv_decimal_ctr_scaling_and_explicit_percentage():
    """Verify decimal CTRs (e.g. 0.0353 -> 3.53%) and explicit percentage strings (0.85% -> 0.85)."""
    csv_content = """campaign_name,spend,clicks,conversions,roas,ctr
Decimal CTR Camp,10000,353,50,2.5,0.0353
Percent CTR Camp,10000,85,20,1.8,0.85%
Standard CTR Camp,10000,500,60,3.2,5.0
"""
    snapshot = parse_csv(csv_content)
    assert len(snapshot.campaigns) == 3
    assert snapshot.campaigns[0].ctr == 3.53
    assert snapshot.campaigns[1].ctr == 0.85
    assert snapshot.campaigns[2].ctr == 5.0


def test_parse_csv_intelligent_campaign_name_synthesis():
    """Verify campaign names and slug IDs are synthesized when campaign_name is missing."""
    csv_content = """platform,campaign_type,industry,country,impressions,clicks,cpc,conversions,revenue
TikTok Ads,Search,EdTech,UK,100000,4000,1.25,200,15000.0
Google Ads,Video,Fintech,UAE,50000,2000,2.00,100,12000.0
Meta Ads,Shopping,SaaS,Germany,80000,3000,1.50,150,18000.0
"""
    snapshot = parse_csv(csv_content)
    assert len(snapshot.campaigns) == 3

    c0 = snapshot.campaigns[0]
    assert c0.platform == Platform.TIKTOK_ADS
    assert c0.campaign_name == "[TikTok Ads] Search - EdTech (UK)"
    assert c0.campaign_id == "cmp_tiktok_ads_search_edtech_uk_1"
    assert c0.spend_inr == 5000.0  # 4000 * 1.25
    assert c0.roas == 3.0  # 15000 / 5000

    c1 = snapshot.campaigns[1]
    assert c1.platform == Platform.GOOGLE_ADS
    assert c1.campaign_name == "[Google Ads] Video - Fintech (UAE)"
    assert c1.campaign_id == "cmp_google_ads_video_fintech_uae_2"

    c2 = snapshot.campaigns[2]
    assert c2.platform == Platform.META_ADS
    assert c2.campaign_name == "[Meta Ads] Shopping - SaaS (Germany)"
    assert c2.campaign_id == "cmp_meta_ads_shopping_saas_germany_3"


def test_parse_json_parity_with_csv_derivations():
    """Verify JSON parsing implements identical metric derivations and name synthesis as CSV."""
    from modules.ads.byod_importer import parse_json

    json_payload = [
        {
            "platform": "TikTok Ads",
            "campaign_type": "Search",
            "industry": "EdTech",
            "country": "UK",
            "clicks": 1000,
            "cpc": 2.0,
            "conversions": 50,
            "revenue": 6000.0,
            "ctr": 0.045,
            "impressions": 22222,
        }
    ]
    snapshot = parse_json(json_payload)
    assert len(snapshot.campaigns) == 1
    c = snapshot.campaigns[0]
    assert c.platform == Platform.TIKTOK_ADS
    assert c.campaign_name == "[TikTok Ads] Search - EdTech (UK)"
    assert c.campaign_id == "cmp_tiktok_ads_search_edtech_uk_1"
    assert c.spend_inr == 2000.0
    assert c.roas == 3.0
    assert c.cpa_inr == 40.0
    assert c.ctr == 4.5


def test_parse_300_plus_sample_multichannel_campaigns_file():
    """Verify sample_multichannel_campaigns.csv contains 300+ coherent rows and parses cleanly."""
    csv_file = Path("services/api/data/sample_multichannel_campaigns.csv")
    assert csv_file.is_file()
    snapshot = parse_csv(csv_file)

    assert len(snapshot.campaigns) >= 300
    assert snapshot.total_spend_inr > 0
    assert snapshot.blended_roas > 0

    platforms = {c.platform for c in snapshot.campaigns}
    assert Platform.GOOGLE_ADS in platforms
    assert Platform.META_ADS in platforms
    assert Platform.TIKTOK_ADS in platforms

    # Ensure all campaign IDs are unique
    camp_ids = [c.campaign_id for c in snapshot.campaigns]
    assert len(camp_ids) == len(set(camp_ids))

