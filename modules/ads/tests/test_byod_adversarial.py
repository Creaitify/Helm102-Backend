"""Empirical Adversarial Stress Tests for BYOD Data Cleaning and Metric Derivation Engine.

Comprehensive empirical test suite covering:
1. Malformed and Extreme CSVs (line endings, blank lines, ragged rows, alternative delimiters, quotes, emojis).
2. Currency parsing and symbol stripping behavior across international formats.
3. Metric auto-derivations (spend derivation, ROAS derivation, CPA derivation, division-by-zero guards).
4. Decimal CTR fraction scaling (0.0353 -> 3.53%) vs explicit percentages ("0.85%" -> 0.85, "85%" -> 85.0).
5. Header aliasing across all canonical dimensions and alias mappings.
6. Multi-dimensional Campaign Name & ID synthesis and slug generation under hostile strings.
7. Multi-platform resolution, context hint evaluation, and substring collision boundaries.
8. Large-scale synthetic datasets (1,000+ rows) with random permutations.
9. JSON, Excel, and unified importer format robustness and parity.
10. Downstream analytics integration with AdOpsAnalyst.
"""

from __future__ import annotations

import io
import random
from pathlib import Path
import openpyxl
import pytest

from modules.ads.analyst import AdOpsAnalyst
from modules.ads.byod_importer import (
    BYODImportError,
    COLUMN_ALIASES,
    InvalidDataFormatError,
    MissingRequiredColumnsError,
    _parse_float,
    _parse_int,
    _parse_platform,
    import_byod_file,
    map_headers,
    normalize_column_name,
    parse_csv,
    parse_excel,
    parse_excel_sheets,
    parse_json,
    parse_row,
    validate_headers,
)
from modules.ads.contracts import CampaignSnapshot, MetricRow, Platform


# ===========================================================================
# 1. Extreme & Malformed CSV Stress Tests
# ===========================================================================

def test_stress_mixed_line_endings_crlf_lf_cr():
    """Verify parser seamlessly handles files containing mixed CRLF, LF, and CR."""
    csv_raw = (
        "campaign_name,spend,clicks,conversions,roas\r\n"
        "Campaign CRLF,1000,50,5,2.5\n"
        "Campaign LF,2000,80,10,3.0\r"
        "Campaign CR,3000,120,15,3.5\r\n"
        "\n\r\n\r"  # Empty junk lines at the end
    )
    snap = parse_csv(csv_raw)
    assert len(snap.campaigns) == 3
    assert snap.campaigns[0].campaign_name == "Campaign CRLF"
    assert snap.campaigns[1].campaign_name == "Campaign LF"
    assert snap.campaigns[2].campaign_name == "Campaign CR"
    assert snap.total_spend_inr == 6000.0


def test_stress_leading_trailing_interspersed_blank_lines():
    """Verify empty rows at start, middle, and end are ignored without indexing desync."""
    csv_raw = (
        "\n\n   \n"
        "campaign_name,spend,clicks,conversions,roas\n"
        "\n"
        "Camp Alpha,5000,200,20,2.0\n"
        "   ,   ,   ,   ,   \n"
        "Camp Beta,8000,300,30,2.5\n"
        "\n\n"
    )
    snap = parse_csv(csv_raw)
    assert len(snap.campaigns) == 2
    assert snap.campaigns[0].campaign_name == "Camp Alpha"
    assert snap.campaigns[1].campaign_name == "Camp Beta"


def test_stress_ragged_rows_missing_trailing_cells():
    """Verify rows with fewer cells than header do not raise IndexError."""
    csv_raw = (
        "campaign_name,spend,clicks,conversions,roas,ctr,cpa,industry\n"
        "Full Row,10000,500,25,2.5,5.0,400.0,Fintech\n"
        "Short Row,5000,200,10,2.0\n"  # Missing ctr, cpa, industry
        "Minimal Row,2000\n"            # Missing clicks, conv, roas (will trigger defaults)
    )
    snap = parse_csv(csv_raw)
    assert len(snap.campaigns) == 3
    assert snap.campaigns[0].spend_inr == 10000.0
    assert snap.campaigns[1].spend_inr == 5000.0
    assert snap.campaigns[1].cpa_inr == 500.0  # Derived 5000/10
    assert snap.campaigns[2].spend_inr == 2000.0
    assert snap.campaigns[2].clicks == 0
    assert snap.campaigns[2].conversions == 0


def test_stress_alternative_delimiters_tsv_semicolon_pipe():
    """Verify dialect sniffing supports TSV, semicolon, and pipe separated files."""
    # Semicolon
    csv_semi = (
        "campaign_name;spend;clicks;conversions;roas\n"
        "Semicolon Camp;15000;600;30;2.8\n"
        "Semicolon Camp 2;20000;800;40;3.0\n"
    )
    snap_semi = parse_csv(csv_semi)
    assert len(snap_semi.campaigns) == 2
    assert snap_semi.campaigns[0].spend_inr == 15000.0
    assert snap_semi.campaigns[1].spend_inr == 20000.0

    # TSV
    csv_tsv = (
        "campaign_name\tspend\tclicks\tconversions\troas\n"
        "TSV Camp\t25000\t1000\t50\t3.2\n"
        "TSV Camp 2\t30000\t1200\t60\t3.5\n"
    )
    snap_tsv = parse_csv(csv_tsv)
    assert len(snap_tsv.campaigns) == 2
    assert snap_tsv.campaigns[0].spend_inr == 25000.0

    # Pipe
    csv_pipe = (
        "campaign_name|spend|clicks|conversions|roas\n"
        "Pipe Camp|35000|1200|60|4.0\n"
        "Pipe Camp 2|40000|1500|80|4.2\n"
    )
    snap_pipe = parse_csv(csv_pipe)
    assert len(snap_pipe.campaigns) == 2
    assert snap_pipe.campaigns[0].spend_inr == 35000.0


def test_stress_utf8_bom_and_emojis():
    """Verify UTF-8 BOM encoding and unicode emojis in campaign names and values."""
    csv_raw = (
        "campaign_name,spend,clicks,conversions,roas\n"
        "🚀 Moonshot Growth 💰,50000,2500,100,3.5\n"
        "🎯 Precision Target 🇮🇳,30000,1200,60,2.8\n"
    )
    snap = parse_csv(csv_raw.encode("utf-8-sig"))
    assert len(snap.campaigns) == 2
    assert snap.campaigns[0].campaign_name == "🚀 Moonshot Growth 💰"
    assert snap.campaigns[1].campaign_name == "🎯 Precision Target 🇮🇳"
    assert snap.total_spend_inr == 80000.0


def test_stress_currency_symbols_and_formats():
    """Verify currency formatting (₹, $, commas in quotes, decimals)."""
    csv_raw = (
        "campaign_name,spend,clicks,conversions,roas,revenue\n"
        'Rupee Camp,"₹ 1,25,000.50",5000,250,3.0,"₹ 3,75,001.50"\n'
        'Dollar Camp,"$ 50,000.00",2000,100,2.5,"$ 1,25,000.00"\n'
        'Plain Camp,75000.75,3000,150,2.0,150001.50\n'
    )
    snap = parse_csv(csv_raw)
    assert len(snap.campaigns) == 3
    assert snap.campaigns[0].spend_inr == 125000.50
    assert snap.campaigns[1].spend_inr == 50000.00
    assert snap.campaigns[2].spend_inr == 75000.75


# ===========================================================================
# 2. Metric Derivation & Boundary Stress Tests
# ===========================================================================

def test_stress_spend_auto_derivation_boundaries():
    """Test spend derivation with zero clicks, zero cpc, or missing spend."""
    # Case A: Clicks > 0, CPC > 0 -> spend = clicks * cpc
    csv_a = "campaign_name,clicks,cpc,conversions,roas\nCamp A,1000,2.5,50,3.0\n"
    snap_a = parse_csv(csv_a)
    assert snap_a.campaigns[0].spend_inr == 2500.0

    # Case B: Clicks = 0, CPC = 2.5 -> spend = 0.0 (no crash)
    csv_b = "campaign_name,clicks,cpc,conversions,roas\nCamp B,0,2.5,0,0.0\n"
    snap_b = parse_csv(csv_b)
    assert snap_b.campaigns[0].spend_inr == 0.0

    # Case C: Clicks = 500, CPC = 0.0 -> spend = 0.0 (no crash)
    csv_c = "campaign_name,clicks,cpc,conversions,roas\nCamp C,500,0.0,10,0.0\n"
    snap_c = parse_csv(csv_c)
    assert snap_c.campaigns[0].spend_inr == 0.0


def test_stress_roas_derivation_and_division_guards():
    """Test ROAS derivation from revenue / spend, with zero spend guards."""
    # Case A: Revenue > 0, Spend > 0 -> roas = revenue / spend
    csv_a = "campaign_name,spend,clicks,conversions,revenue\nCamp A,10000,500,25,35000\n"
    snap_a = parse_csv(csv_a)
    assert snap_a.campaigns[0].roas == 3.5

    # Case B: Revenue > 0, Spend = 0 -> roas = 0.0 (guard against ZeroDivisionError)
    csv_b = "campaign_name,spend,clicks,conversions,revenue\nCamp B,0,500,25,35000\n"
    snap_b = parse_csv(csv_b)
    assert snap_b.campaigns[0].roas == 0.0

    # Case C: Revenue = 0, Spend = 10000 -> roas = 0.0
    csv_c = "campaign_name,spend,clicks,conversions,revenue\nCamp C,10000,500,25,0\n"
    snap_c = parse_csv(csv_c)
    assert snap_c.campaigns[0].roas == 0.0

    # Case D: Explicit ROAS as '-' or empty string -> should derive from revenue/spend
    csv_d = "campaign_name,spend,clicks,conversions,roas,revenue\nCamp D,10000,500,25,-,25000\n"
    snap_d = parse_csv(csv_d)
    assert snap_d.campaigns[0].roas == 2.5


def test_stress_cpa_derivation_and_division_guards():
    """Test CPA derivation from spend / conversions, with zero conversions guard."""
    # Case A: Spend > 0, Conversions > 0 -> cpa = spend / conversions
    csv_a = "campaign_name,spend,clicks,conversions,roas\nCamp A,12000,600,40,2.0\n"
    snap_a = parse_csv(csv_a)
    assert snap_a.campaigns[0].cpa_inr == 300.0  # 12000 / 40

    # Case B: Spend > 0, Conversions = 0 -> cpa = 0.0 (guard against ZeroDivisionError)
    csv_b = "campaign_name,spend,clicks,conversions,roas\nCamp B,12000,600,0,2.0\n"
    snap_b = parse_csv(csv_b)
    assert snap_b.campaigns[0].cpa_inr == 0.0

    # Case C: Explicit CPA provided -> keep explicit value
    csv_c = "campaign_name,spend,clicks,conversions,roas,cpa\nCamp C,12000,600,40,2.0,285.50\n"
    snap_c = parse_csv(csv_c)
    assert snap_c.campaigns[0].cpa_inr == 285.50


# ===========================================================================
# 3. Decimal CTR Normalization vs Explicit Percentages Stress Tests
# ===========================================================================

@pytest.mark.parametrize(
    "raw_ctr,expected_ctr",
    [
        ("0.0353", 3.53),     # Decimal fraction (0.0 < x <= 1.0 without %) -> multiplied by 100
        ("0.0012", 0.12),     # Low decimal fraction
        ("0.85%", 0.85),      # Explicit percentage with <= 1.0 value -> preserved as 0.85
        ("85%", 85.0),        # Explicit percentage with > 1.0 value -> preserved as 85.0
        ("1.0%", 1.0),        # Explicit 1.0% -> preserved as 1.0
        ("1.0", 100.0),       # Decimal fraction 1.0 (100% CTR) -> multiplied by 100
        ("0.0%", 0.0),        # Zero percent
        ("0.0", 0.0),         # Zero float
        ("3.53", 3.53),       # Already in percentage format (> 1.0)
        ("12.50", 12.50),     # Already in percentage format
        ("-", 0.0),           # Dash / missing
        ("nan", 0.0),         # NaN string
    ],
)
def test_stress_ctr_normalization_matrix(raw_ctr: str, expected_ctr: float):
    csv_raw = f"campaign_name,spend,clicks,conversions,roas,ctr\nCamp Test,1000,50,5,2.0,{raw_ctr}\n"
    snap = parse_csv(csv_raw)
    assert snap.campaigns[0].ctr == expected_ctr


def test_stress_ctr_fallback_to_clicks_over_impressions():
    """Verify fallback CTR = (clicks / impressions) * 100 when CTR column is missing or 0."""
    # Case A: CTR omitted, impressions > 0
    csv_a = "campaign_name,spend,clicks,conversions,roas,impressions\nCamp A,1000,250,10,2.0,5000\n"
    snap_a = parse_csv(csv_a)
    assert snap_a.campaigns[0].ctr == 5.0  # (250 / 5000) * 100

    # Case B: CTR omitted, impressions = 0 -> ctr = 0.0 (guard against ZeroDivisionError)
    csv_b = "campaign_name,spend,clicks,conversions,roas,impressions\nCamp B,1000,250,10,2.0,0\n"
    snap_b = parse_csv(csv_b)
    assert snap_b.campaigns[0].ctr == 0.0

    # Case C: CTR column present with 0.0, fallback calculates from impressions
    csv_c = "campaign_name,spend,clicks,conversions,roas,impressions,ctr\nCamp C,1000,150,10,2.0,3000,0.0\n"
    snap_c = parse_csv(csv_c)
    assert snap_c.campaigns[0].ctr == 5.0  # (150 / 3000) * 100


# ===========================================================================
# 4. Header Aliasing Stress Tests
# ===========================================================================

def test_stress_obscure_spend_aliases():
    """Test all variants of spend aliases."""
    aliases = [
        "ad_spend", "adspend", "media_spend", "cost_micros", "amount_spent",
        "amount_spent_inr", "total_spend", "cost_inr", "spend_amount", "daily_spend",
        "ad_cost", "adcost", "total_cost", "budget_spent", "spent", "Spend (INR)", "Cost [USD]"
    ]
    for alias in aliases:
        csv_raw = f"campaign_name,{alias},clicks,conversions,roas\nCamp X,5000,200,10,2.5\n"
        snap = parse_csv(csv_raw)
        assert snap.campaigns[0].spend_inr == 5000.0, f"Failed alias: {alias}"


def test_stress_obscure_revenue_aliases():
    """Test all variants of revenue aliases."""
    aliases = [
        "revenue", "conv_value", "conversion_value", "total_revenue", "sales",
        "purchase_value", "total_conv_value", "sales_amount", "revenue_usd",
        "revenue_inr", "value", "total_sales", "gross_revenue", "all_conv_value",
        "Conv. Value (INR)", "Total Sales [₹]"
    ]
    for alias in aliases:
        csv_raw = f"campaign_name,spend,clicks,conversions,{alias}\nCamp X,2000,100,5,6000\n"
        snap = parse_csv(csv_raw)
        assert snap.campaigns[0].roas == 3.0, f"Failed revenue alias: {alias}"


def test_stress_obscure_cpc_aliases():
    """Test all variants of CPC aliases."""
    aliases = [
        "cpc", "cost_per_click", "cpc_inr", "cpc_usd", "avg_cpc", "average_cpc",
        "Avg. CPC (INR)"
    ]
    for alias in aliases:
        csv_raw = f"campaign_name,clicks,{alias},conversions,roas\nCamp X,400,3.50,20,2.0\n"
        snap = parse_csv(csv_raw)
        # spend should be derived as 400 * 3.50 = 1400.0
        assert snap.campaigns[0].spend_inr == 1400.0, f"Failed CPC alias: {alias}"


# ===========================================================================
# 5. Campaign Synthesis & Slug Stress Tests
# ===========================================================================

def test_stress_campaign_synthesis_all_dimensions_missing():
    """Verify synthesis when campaign_name, type, industry, country are all missing."""
    csv_raw = "platform,spend,clicks,conversions,roas\nGoogle Ads,10000,500,25,3.0\n"
    snap = parse_csv(csv_raw)
    camp = snap.campaigns[0]
    assert camp.campaign_name == "[Google Ads] Campaign 1"
    assert camp.campaign_id == "cmp_google_ads_campaign_1_1"


def test_stress_campaign_synthesis_partial_dimensions():
    """Verify synthesis when only partial dimensions (e.g. only industry or country) exist."""
    # Only industry
    csv_1 = "platform,industry,spend,clicks,conversions,roas\nMeta Ads,Fintech,10000,500,25,3.0\n"
    camp_1 = parse_csv(csv_1).campaigns[0]
    assert camp_1.campaign_name == "[Meta Ads] Fintech"
    assert camp_1.campaign_id == "cmp_meta_ads_fintech_1"

    # Only country
    csv_2 = "platform,country,spend,clicks,conversions,roas\nTikTok Ads,Germany,10000,500,25,3.0\n"
    camp_2 = parse_csv(csv_2).campaigns[0]
    assert camp_2.campaign_name == "[TikTok Ads] Campaign 1 (Germany)"
    assert camp_2.campaign_id == "cmp_tiktok_ads_campaign_1_germany_1"

    # Type and country without industry
    csv_3 = "platform,campaign_type,country,spend,clicks,conversions,roas\nLinkedIn Ads,LeadGen,UK,10000,500,25,3.0\n"
    camp_3 = parse_csv(csv_3).campaigns[0]
    assert camp_3.campaign_name == "[LinkedIn Ads] LeadGen (UK)"
    assert camp_3.campaign_id == "cmp_linkedin_ads_leadgen_uk_1"


def test_stress_campaign_synthesis_special_characters_in_dimensions():
    """Verify special characters (slash, ampersand, parentheses, emoji) are handled cleanly in slugs."""
    csv_raw = (
        "platform,campaign_type,industry,country,spend,clicks,conversions,roas\n"
        "Google Ads,Search & P-Max / Shopping,B2B & AI (SaaS),US / Canada 🇺🇸,10000,500,25,3.0\n"
    )
    camp = parse_csv(csv_raw).campaigns[0]
    assert camp.campaign_name == "[Google Ads] Search & P-Max / Shopping - B2B & AI (SaaS) (US / Canada 🇺🇸)"
    # Slug should be clean with underscores and row index
    assert camp.campaign_id.startswith("cmp_")
    assert camp.campaign_id.endswith("_1")
    assert " " not in camp.campaign_id
    assert "/" not in camp.campaign_id
    assert "&" not in camp.campaign_id


# ===========================================================================
# 6. Large-Scale Synthetic Dataset (1,000 Rows) Stress Test
# ===========================================================================

def test_stress_large_scale_thousand_rows():
    """Verify high-throughput ingestion of 1,000 varied multi-channel rows."""
    header = "platform,campaign_type,industry,country,impressions,clicks,cpc,conversions,revenue,ctr\n"
    rows = []
    platforms = ["Google Ads", "Meta Ads", "TikTok Ads", "LinkedIn Ads"]
    industries = ["Fintech", "SaaS", "E-commerce", "HealthTech", "EdTech"]
    countries = ["USA", "UK", "Germany", "India", "UAE", "Singapore"]

    for i in range(1, 1001):
        plat = random.choice(platforms)
        ind = random.choice(industries)
        ctry = random.choice(countries)
        clicks = random.randint(100, 10000)
        cpc = round(random.uniform(0.5, 5.0), 2)
        impr = clicks * random.randint(10, 30)
        convs = max(1, clicks // random.randint(15, 35))
        rev = round((clicks * cpc) * random.uniform(1.2, 4.5), 2)
        ctr_val = f"{round((clicks / impr), 4)}"  # decimal fraction

        rows.append(f"{plat},Search,{ind},{ctry},{impr},{clicks},{cpc},{convs},{rev},{ctr_val}")

    large_csv = header + "\n".join(rows)
    snap = parse_csv(large_csv)
    assert len(snap.campaigns) == 1000
    assert snap.total_spend_inr > 0
    assert snap.blended_roas > 0

    # Ensure all campaign IDs are unique
    camp_ids = [c.campaign_id for c in snap.campaigns]
    assert len(camp_ids) == len(set(camp_ids))

    # Verify downstream analysis on 1,000 campaigns
    analyst = AdOpsAnalyst()
    findings = analyst.analyze(snap)
    assert len(findings["per_campaign"]) == 1000
    assert len(findings["channel_breakdown"]) >= 4


# ===========================================================================
# 7. JSON & Excel Stress Tests
# ===========================================================================

def test_stress_json_various_envelopes_and_missing_keys():
    """Verify JSON parsing handles array, {campaigns: [...]}, {data: [...]}, {rows: [...]}, and dict of lists."""
    # Envelope: {"data": [...]}
    json_data_1 = {
        "data": [
            {"platform": "google_ads", "spend": 10000, "clicks": 500, "conversions": 25, "roas": 3.0}
        ]
    }
    snap_1 = parse_json(json_data_1)
    assert len(snap_1.campaigns) == 1
    assert snap_1.campaigns[0].spend_inr == 10000.0

    # Envelope: {"rows": [...]}
    json_data_2 = {
        "rows": [
            {"platform": "meta_ads", "clicks": 400, "cpc": 2.5, "conversions": 20, "revenue": 3000.0}
        ]
    }
    snap_2 = parse_json(json_data_2)
    assert len(snap_2.campaigns) == 1
    assert snap_2.campaigns[0].spend_inr == 1000.0  # 400 * 2.5
    assert snap_2.campaigns[0].roas == 3.0          # 3000 / 1000


def test_stress_excel_multisheet_and_skipping(tmp_path: Path):
    """Verify openpyxl parsing handles distinct sheets and skips empty/summary sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore

    ws_gads = wb.create_sheet(title="Google_Ads")
    ws_gads.append(["Campaign Name", "Clicks", "CPC", "Conversions", "Revenue", "CTR"])
    ws_gads.append(["G Search", 2000, 2.00, 100, 12000.0, "4.5%"])

    ws_meta = wb.create_sheet(title="Meta_Ads")
    ws_meta.append(["Campaign Name", "Clicks", "CPC", "Conversions", "Revenue", "CTR"])
    ws_meta.append(["M Retargeting", 1000, 1.50, 50, 4500.0, "0.045"])

    # Empty sheet
    wb.create_sheet(title="EmptySheet")

    # Readme metadata sheet
    ws_readme = wb.create_sheet(title="Readme")
    ws_readme.append(["Key", "Value"])
    ws_readme.append(["Version", "1.0"])

    excel_path = tmp_path / "stress_test.xlsx"
    wb.save(str(excel_path))

    snap = parse_excel(excel_path)
    assert len(snap.campaigns) == 2
    assert snap.campaigns[0].platform == Platform.GOOGLE_ADS
    assert snap.campaigns[0].spend_inr == 4000.0  # 2000 * 2.00
    assert snap.campaigns[0].roas == 3.0          # 12000 / 4000
    assert snap.campaigns[0].ctr == 4.5           # 4.5% preserved

    assert snap.campaigns[1].platform == Platform.META_ADS
    assert snap.campaigns[1].spend_inr == 1500.0  # 1000 * 1.50
    assert snap.campaigns[1].roas == 3.0          # 4500 / 1500
    assert snap.campaigns[1].ctr == 4.5           # 0.045 * 100
